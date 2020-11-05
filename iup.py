# -*- coding: utf-8 -*-
"""
Created on Thu Nov  5 17:45:22 2020

@author: Марквирер Владлена Дмитриевна
"""
import time
start_time = time.time()
import locale, os, re
import openpyxl as excel
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font

locale.setlocale(locale.LC_ALL, "") 

wb_up = excel.load_workbook(filename = 'Агишев.xlsx')
sheet_up = wb_up['Учебный план студента']

ar = sheet_up['K1'].value
star = sheet_up['B34'].value
data = sheet_up['B35'].value
osup = sheet_up['B38'].value
name = sheet_up['C39'].value

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

groups = ['ПИ-18', 'ПИ-19']

for group in groups:
    directory ='./'+group
    students = os.listdir(directory)
    for student in students:
        wb_up_corr = excel.load_workbook(filename = './' + group + '/'+ student)
        sheet_up_corr = wb_up_corr['Учебный план студента']
        
        row_count = sheet_up_corr.max_row
        row_count_str = str(row_count)
        date_row = str(row_count - 6)
        star_row = str(row_count - 7)
        delete_row1 = str(row_count - 3)
        delete_row2 = str(row_count - 4)
        osup_row = delete_row1
        name_row = str(row_count - 2)
        delete_row3 = str(row_count - 1)
        delete_row4 = str(row_count)
        itogo_row = str(row_count - 8)
    
        # И.о. академического руководителя
        sheet_up_corr['K1'].value = ar
        # Дата
        sheet_up_corr['B' + date_row].value = data 
        # звёздочка
        sheet_up_corr['B' + star_row].value = star
        # удалить строку
        sheet_up_corr['B' + delete_row1].value = "" 
        sheet_up_corr['B' + delete_row2].value = ""
        sheet_up_corr['B' + delete_row3].value = ""
        sheet_up_corr['C' + delete_row4].value = ""
        # ОСУП
        sheet_up_corr['B' + osup_row].value = osup
        sheet_up_corr['B' + osup_row].alignment = Alignment(horizontal="left")
        sheet_up_corr['B' + osup_row].font = Font(bold = True)
        # подпись и расшифровка
        sheet_up_corr['C' + name_row].value = name
        sheet_up_corr.merge_cells('C'+name_row + ':D'+name_row)
        
        diapasons = ['B'+itogo_row + ':C'+itogo_row, 'A14:A16', 'B14:B16', 'C14:C16', 'D14:D16', 'G14:G16', 'H14:H16', 'I14:I16',
             'E14:F15', 'J14:M15']
        
        for diapason in diapasons:
            sheet_up_corr.merge_cells(diapason)
            d1 = diapason[:diapason.index(':')]
            d2 = diapason[diapason.index(':')+1:]
            cStart = d1[:re.search("\d", d1).start()] 
            cStop = d2[:re.search("\d", d2).start()]
            rStart = d1[re.search("\d", d1).start():]
            rStop = d2[re.search("\d", d1).start():]
            # только для столбцов с одной буквой
            for i in range(ord(cStop)-ord(cStart)+1):
                for j in range(int(rStop)-int(rStart) + 1):
                   sheet_up_corr[chr(ord(cStart)+i)+str(int(rStart)+j)].border = thin_border
                
        # сохранение в отдельный файл
        wb_up_corr.save('./Результат/'+ group +'/'+student)
print("--- %s seconds ---" % (time.time() - start_time))